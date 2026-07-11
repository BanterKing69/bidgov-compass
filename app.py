"""BidGov Compass — Flask backend for the tender aggregator dashboard."""

from __future__ import annotations

import io
import json
import subprocess
import sys
import threading
import time
from datetime import datetime, timezone
from pathlib import Path

from flask import (
    Flask, Response, jsonify, render_template, request, send_file, abort
)
from flask_login import login_required, current_user

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parent / "collectors"))

import db  # noqa: E402
import chat as chat_mod  # noqa: E402
import auth  # noqa: E402

APP_ROOT = Path(__file__).resolve().parent
app = Flask(__name__, template_folder="templates", static_folder="static")
auth.init_app(app)

# --------------------------------------------------------------------------- #
# Scrape job state (in-memory; one job at a time)
# --------------------------------------------------------------------------- #
_scrape_lock = threading.Lock()
_scrape_state: dict = {
    "running": False, "started_at": None, "finished_at": None,
    "days_back": None, "log": [], "error": None,
    "before": {}, "after": {},
}


def _snapshot() -> dict:
    conn = db.connect()
    try:
        return db.stats(conn)
    finally:
        conn.close()


def _run_scrape(days_back: int, max_pages: int) -> None:
    global _scrape_state
    _scrape_state["log"] = []
    _scrape_state["error"] = None
    _scrape_state["before"] = _snapshot()
    _scrape_state["started_at"] = datetime.now(timezone.utc).isoformat()

    try:
        proc = subprocess.Popen(
            [sys.executable, str(APP_ROOT / "run.py"),
             "--days-back", str(days_back), "--max-pages", str(max_pages)],
            cwd=str(APP_ROOT), stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT, text=True, bufsize=1,
        )
        assert proc.stdout is not None
        for line in proc.stdout:
            _scrape_state["log"].append(line.rstrip())
            _scrape_state["log"] = _scrape_state["log"][-500:]
        proc.wait()
        if proc.returncode != 0:
            _scrape_state["error"] = f"exit code {proc.returncode}"
    except Exception as exc:  # noqa: BLE001
        _scrape_state["error"] = str(exc)
    finally:
        _scrape_state["after"] = _snapshot()
        _scrape_state["finished_at"] = datetime.now(timezone.utc).isoformat()
        _scrape_state["running"] = False


# --------------------------------------------------------------------------- #
# Auth gate: every route except the auth blueprint, static files, and /health
# requires login. API endpoints return JSON 401; page routes redirect to login.
# --------------------------------------------------------------------------- #
_PUBLIC_PATHS = ("/auth/", "/static/", "/health")


@app.before_request
def _require_login():
    path = request.path or "/"
    if any(path == p or path.startswith(p) for p in _PUBLIC_PATHS):
        return None
    if current_user.is_authenticated:
        return None
    # JSON responses for API paths, redirect for page routes
    if path.startswith("/api/"):
        return jsonify({"error": "auth_required"}), 401
    from flask import redirect, url_for
    return redirect(url_for("auth.login", next=path))


# --------------------------------------------------------------------------- #
# Routes
# --------------------------------------------------------------------------- #
@app.route("/")
def home():
    # / is the Search screen (full explorer). Endpoint name preserved so the
    # existing url_for('home') references (auth redirects, templates) keep
    # resolving.
    return render_template("dashboard.html", user=current_user)


@app.route("/live-bids")
def live_bids():
    # Live bids page — client-facing, urgency-forward view of open tenders.
    # Server-side locks deadline >= now on every /api/live-tenders call;
    # this page just renders the shell.
    return render_template("live_bids.html", user=current_user)


@app.route("/dashboard")
def dashboard_page():
    # Dashboard page — KPI row + the five Chart.js charts, moved off /.
    return render_template("dashboard_page.html", user=current_user)


# --- filters + table ----------------------------------------------------- #
_ALLOWED_SORT = {
    # deadline
    "deadline":       "(deadline IS NULL), deadline ASC",
    "deadline_desc":  "(deadline IS NULL), deadline DESC",
    # value
    "value_desc":     "(value_amount IS NULL), value_amount DESC",
    "value_asc":      "(value_amount IS NULL), value_amount ASC",
    # published
    "published_desc": "(published_date IS NULL), published_date DESC",
    "published_asc":  "(published_date IS NULL), published_date ASC",
    # title / buyer / category / source
    "title_asc":      "title COLLATE NOCASE ASC",
    "title_desc":     "title COLLATE NOCASE DESC",
    "buyer_asc":      "(buyer_name IS NULL), buyer_name COLLATE NOCASE ASC",
    "buyer_desc":     "(buyer_name IS NULL), buyer_name COLLATE NOCASE DESC",
    "category_asc":   "(category IS NULL), category COLLATE NOCASE ASC",
    "category_desc":  "(category IS NULL), category COLLATE NOCASE DESC",
    "source_asc":     "source COLLATE NOCASE ASC",
    "source_desc":    "source COLLATE NOCASE DESC",
}

# Awards-only sort keys (over the same tenders table, but different columns)
_ALLOWED_AWARD_SORT = {
    "awarded_date":       "(awarded_date IS NULL), awarded_date DESC",   # newest first (default)
    "awarded_date_asc":   "(awarded_date IS NULL), awarded_date ASC",
    "awarded_value_desc": "(awarded_value_amount IS NULL), awarded_value_amount DESC",
    "awarded_value_asc":  "(awarded_value_amount IS NULL), awarded_value_amount ASC",
    "contract_end_asc":   "(contract_end_date IS NULL), contract_end_date ASC",
    "contract_end_desc":  "(contract_end_date IS NULL), contract_end_date DESC",
    "supplier_asc":       "(awarded_supplier_name IS NULL), awarded_supplier_name COLLATE NOCASE ASC",
    "supplier_desc":      "(awarded_supplier_name IS NULL), awarded_supplier_name COLLATE NOCASE DESC",
    "title_asc":          "title COLLATE NOCASE ASC",
    "title_desc":         "title COLLATE NOCASE DESC",
    "buyer_asc":          "(buyer_name IS NULL), buyer_name COLLATE NOCASE ASC",
    "buyer_desc":         "(buyer_name IS NULL), buyer_name COLLATE NOCASE DESC",
    "category_asc":       "(category IS NULL), category COLLATE NOCASE ASC",
    "category_desc":      "(category IS NULL), category COLLATE NOCASE DESC",
    "source_asc":         "source COLLATE NOCASE ASC",
    "source_desc":        "source COLLATE NOCASE DESC",
}


def _build_where(args, *, stage: str = "tender") -> tuple[str, list]:
    """Build a WHERE clause for /api/tenders or /api/awards.

    stage='tender'  -> notice_stage='tender' (default; the opportunities view)
    stage='award'   -> notice_stage='award'  (the won-contracts view)
    stage='any'     -> no stage restriction (used by chat / raw pivot)
    """
    clauses, params = [], []

    if stage == "tender":
        clauses.append("notice_stage = 'tender'")
    elif stage == "award":
        clauses.append("notice_stage = 'award'")
    # stage == "any" -> no stage filter

    if stage == "tender" and args.get("open_only", "1") == "1":
        clauses.append("is_open = 1")

    if stage == "award":
        # Awards-specific filters
        sup = (args.get("supplier") or "").strip()
        if sup:
            clauses.append("LOWER(awarded_supplier_name) LIKE ?")
            params.append(f"%{sup.lower()}%")
        ad_from = args.get("awarded_from")
        ad_to = args.get("awarded_to")
        if ad_from:
            clauses.append("awarded_date >= ?"); params.append(ad_from)
        if ad_to:
            clauses.append("awarded_date <= ?"); params.append(ad_to)
        av_min = args.get("awarded_min", type=float)
        av_max = args.get("awarded_max", type=float)
        if av_min is not None:
            clauses.append("awarded_value_amount >= ?"); params.append(av_min)
        if av_max is not None:
            clauses.append("awarded_value_amount <= ?"); params.append(av_max)

    cats = args.getlist("category")
    if cats:
        clauses.append(f"category IN ({','.join(['?']*len(cats))})")
        params.extend(cats)

    sources = args.getlist("source")
    if sources:
        clauses.append(f"source IN ({','.join(['?']*len(sources))})")
        params.extend(sources)

    vmin = args.get("value_min", type=float)
    vmax = args.get("value_max", type=float)
    if vmin is not None:
        clauses.append("value_amount >= ?"); params.append(vmin)
    if vmax is not None:
        clauses.append("value_amount <= ?"); params.append(vmax)

    if stage != "award":
        dl_before = args.get("deadline_before")
        dl_after = args.get("deadline_after")
        if dl_after:
            clauses.append("deadline >= ?"); params.append(dl_after)
        if dl_before:
            clauses.append("deadline <= ?"); params.append(dl_before)

    q = args.get("q", "").strip()
    if q:
        clauses.append("(LOWER(title) LIKE ? OR LOWER(buyer_name) LIKE ?)")
        term = f"%{q.lower()}%"
        params.extend([term, term])

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    return where, params


@app.route("/api/tenders")
def api_tenders():
    where, params = _build_where(request.args)
    sort = _ALLOWED_SORT.get(request.args.get("sort", "deadline"),
                             _ALLOWED_SORT["deadline"])
    limit = min(int(request.args.get("limit", 500)), 5000)
    cols = ("uid,source,title,category,cpv_code,buyer_name,buyer_region,"
            "value_amount,value_currency,published_date,deadline,status,"
            "is_open,notice_url")
    conn = db.connect()
    try:
        cur = conn.execute(
            f"SELECT {cols} FROM tenders {where} ORDER BY {sort} LIMIT ?",
            [*params, limit],
        )
        col_names = [d[0] for d in cur.description]
        rows = [dict(zip(col_names, r)) for r in cur.fetchall()]
        total_cur = conn.execute(f"SELECT COUNT(*) FROM tenders {where}", params)
        total = total_cur.fetchone()[0]
    finally:
        conn.close()
    return jsonify({"total": total, "returned": len(rows), "rows": rows})


# --------------------------------------------------------------------------- #
# Live-bids API — same table + columns as /api/tenders, but with a hardcoded
# server-side "deadline >= now" clause the client cannot remove. is_open is
# set at collection time and goes stale (verified: DB has ~5 rows where
# is_open=1 but deadline has already passed), so we compare deadline to
# datetime('now') at query time. SQLite compares ISO 8601 strings lexically
# — safe for the well-formed ISO deadlines our normaliser produces.
# --------------------------------------------------------------------------- #
def _build_live_where(args):
    """Same as _build_where(stage='tender') but always AND'd with
    `deadline IS NOT NULL AND deadline >= datetime('now')`. Returns (where, params).
    """
    where, params = _build_where(args)
    # _build_where always emits WHERE ... for stage='tender' (at least the
    # notice_stage clause is present), so we can safely append with AND.
    where = where + " AND deadline IS NOT NULL AND deadline >= datetime('now')"
    return where, params


@app.route("/api/live-tenders")
def api_live_tenders():
    where, params = _build_live_where(request.args)
    sort = _ALLOWED_SORT.get(request.args.get("sort", "deadline"),
                             _ALLOWED_SORT["deadline"])
    limit = min(int(request.args.get("limit", 500)), 5000)
    cols = ("uid,source,title,category,cpv_code,buyer_name,buyer_region,"
            "value_amount,value_currency,published_date,deadline,status,"
            "is_open,notice_url")
    conn = db.connect()
    try:
        cur = conn.execute(
            f"SELECT {cols} FROM tenders {where} ORDER BY {sort} LIMIT ?",
            [*params, limit],
        )
        col_names = [d[0] for d in cur.description]
        rows = [dict(zip(col_names, r)) for r in cur.fetchall()]
        total_cur = conn.execute(f"SELECT COUNT(*) FROM tenders {where}", params)
        total = total_cur.fetchone()[0]
    finally:
        conn.close()
    return jsonify({"total": total, "returned": len(rows), "rows": rows})


@app.route("/api/live-stats")
def api_live_stats():
    """KPI-strip payload for the Live bids page: open count, total open value,
    closing ≤7d count. Honours the same filter params /api/live-tenders accepts."""
    where, params = _build_live_where(request.args)
    conn = db.connect()
    try:
        row = conn.execute(
            f"SELECT COUNT(*), "
            f"       ROUND(COALESCE(SUM(value_amount), 0)), "
            f"       SUM(CASE WHEN deadline <= datetime('now', '+7 days') THEN 1 ELSE 0 END) "
            f"FROM tenders {where}",
            params,
        ).fetchone()
    finally:
        conn.close()
    return jsonify({
        "open_count":       row[0] or 0,
        "total_open_value": row[1] or 0,
        "closing_7d":       row[2] or 0,
    })


# --------------------------------------------------------------------------- #
# Awards ("Won contracts") — parallel routes that scope to notice_stage='award'
# --------------------------------------------------------------------------- #
@app.route("/awards")
def awards_page():
    return render_template("awards.html", user=current_user)


@app.route("/api/awards")
def api_awards():
    where, params = _build_where(request.args, stage="award")
    sort = _ALLOWED_AWARD_SORT.get(
        request.args.get("sort", "awarded_date"),
        _ALLOWED_AWARD_SORT["awarded_date"],
    )
    limit = min(int(request.args.get("limit", 500)), 5000)
    cols = ("uid,source,title,category,cpv_code,buyer_name,buyer_region,"
            "value_amount,awarded_value_amount,awarded_value_currency,"
            "awarded_supplier_name,awarded_supplier_id,awarded_supplier_count,"
            "awarded_date,contract_start_date,contract_end_date,"
            "notice_url,ocid")
    conn = db.connect()
    try:
        cur = conn.execute(
            f"SELECT {cols} FROM tenders {where} ORDER BY {sort} LIMIT ?",
            [*params, limit],
        )
        col_names = [d[0] for d in cur.description]
        rows = [dict(zip(col_names, r)) for r in cur.fetchall()]
        total = conn.execute(f"SELECT COUNT(*) FROM tenders {where}", params).fetchone()[0]
    finally:
        conn.close()
    return jsonify({"total": total, "returned": len(rows), "rows": rows})


@app.route("/api/awards/facets")
def api_awards_facets():
    conn = db.connect()
    try:
        cats = [r[0] for r in conn.execute(
            "SELECT DISTINCT category FROM tenders "
            "WHERE notice_stage='award' AND category IS NOT NULL "
            "ORDER BY category").fetchall()]
        sources = [r[0] for r in conn.execute(
            "SELECT DISTINCT source FROM tenders WHERE notice_stage='award' "
            "ORDER BY source").fetchall()]
        top_suppliers = conn.execute(
            "SELECT awarded_supplier_name, COUNT(*) c, "
            "       ROUND(COALESCE(SUM(awarded_value_amount),0)) v "
            "FROM tenders WHERE notice_stage='award' "
            "AND awarded_supplier_name IS NOT NULL "
            "GROUP BY awarded_supplier_name ORDER BY c DESC LIMIT 100"
        ).fetchall()
        val = conn.execute(
            "SELECT MIN(awarded_value_amount), MAX(awarded_value_amount) "
            "FROM tenders WHERE notice_stage='award' "
            "AND awarded_value_amount IS NOT NULL").fetchone()
        dates = conn.execute(
            "SELECT MIN(awarded_date), MAX(awarded_date) "
            "FROM tenders WHERE notice_stage='award'").fetchone()
    finally:
        conn.close()
    return jsonify({
        "categories": cats,
        "sources": sources,
        "top_suppliers": [
            {"name": s[0], "wins": s[1], "value": s[2]}
            for s in top_suppliers if s[0]
        ],
        "value_min": val[0] or 0,
        "value_max": val[1] or 0,
        "date_min": dates[0],
        "date_max": dates[1],
    })


@app.route("/api/awards/stats")
def api_awards_stats():
    where, params = _build_where(request.args, stage="award")
    conn = db.connect()
    try:
        totals = conn.execute(
            f"SELECT COUNT(*), "
            f"       COUNT(DISTINCT awarded_supplier_name), "
            f"       ROUND(COALESCE(SUM(awarded_value_amount),0)) "
            f"FROM tenders {where}", params).fetchone()

        # Median awarded value — same window-function pattern as tender stats
        median_row = conn.execute(f"""
            WITH ranked AS (
                SELECT awarded_value_amount v, ROW_NUMBER() OVER (ORDER BY awarded_value_amount) rn,
                       COUNT(*) OVER () c
                FROM tenders {where} AND awarded_value_amount IS NOT NULL
                                     AND awarded_value_amount > 0
            )
            SELECT ROUND(AVG(v)) FROM ranked WHERE rn IN ((c+1)/2, (c+2)/2)
        """, params).fetchone()
        median_value = (median_row and median_row[0]) or 0

        by_supplier = conn.execute(
            f"SELECT awarded_supplier_name, COUNT(*) c, "
            f"       ROUND(COALESCE(SUM(awarded_value_amount),0)) v "
            f"FROM tenders {where} AND awarded_supplier_name IS NOT NULL "
            f"GROUP BY awarded_supplier_name ORDER BY c DESC LIMIT 15",
            params).fetchall()

        by_cat = conn.execute(
            f"SELECT COALESCE(category,'(unmapped)') k, COUNT(*) n, "
            f"       ROUND(COALESCE(SUM(awarded_value_amount),0)) total_v "
            f"FROM tenders {where} GROUP BY category ORDER BY n DESC LIMIT 40",
            params).fetchall()

        by_month = conn.execute(
            f"SELECT substr(awarded_date,1,7) k, COUNT(*) n, "
            f"       ROUND(COALESCE(SUM(awarded_value_amount),0)) total_v "
            f"FROM tenders {where} AND awarded_date IS NOT NULL "
            f"GROUP BY 1 ORDER BY 1", params).fetchall()

        by_source = conn.execute(
            f"SELECT source k, COUNT(*) n FROM tenders {where} "
            f"GROUP BY source ORDER BY n DESC", params).fetchall()

        # Upcoming renewal windows (contracts expiring in the next 12 months)
        renewals = conn.execute(f"""
            SELECT substr(contract_end_date,1,7) k, COUNT(*) n
            FROM tenders {where} AND contract_end_date IS NOT NULL
            AND contract_end_date >= date('now')
            AND contract_end_date <  date('now','+365 days')
            GROUP BY 1 ORDER BY 1
        """, params).fetchall()
    finally:
        conn.close()

    def d2(rows, keys=("k", "n")):
        return [dict(zip(keys, r)) for r in rows]

    return jsonify({
        "totals": {
            "awards":          totals[0],
            "unique_suppliers": totals[1] or 0,
            "total_value":     totals[2] or 0,
            "median_value":    median_value,
        },
        "by_supplier":   [{"k": r[0], "n": r[1], "total_v": r[2]} for r in by_supplier],
        "by_category":   [{"k": r[0], "n": r[1], "total_v": r[2]} for r in by_cat],
        "by_month":      [{"k": r[0], "n": r[1], "total_v": r[2]} for r in by_month],
        "by_source":     d2(by_source),
        "renewals":      d2(renewals),
    })


@app.route("/api/facets")
def api_facets():
    conn = db.connect()
    try:
        cats = [r[0] for r in conn.execute(
            "SELECT DISTINCT category FROM tenders "
            "WHERE notice_stage='tender' AND category IS NOT NULL "
            "ORDER BY category").fetchall()]
        sources = [r[0] for r in conn.execute(
            "SELECT DISTINCT source FROM tenders WHERE notice_stage='tender' "
            "ORDER BY source").fetchall()]
        mx = conn.execute(
            "SELECT MIN(value_amount), MAX(value_amount) FROM tenders "
            "WHERE notice_stage='tender' AND value_amount IS NOT NULL").fetchone()
    finally:
        conn.close()
    return jsonify({
        "categories": cats, "sources": sources,
        "value_min": mx[0] or 0, "value_max": mx[1] or 0,
    })


# --- stats for charts ---------------------------------------------------- #
@app.route("/api/stats")
def api_stats():
    where, params = _build_where(request.args)
    conn = db.connect()
    try:
        def q(sql):
            return conn.execute(sql, params).fetchall()

        by_cat = q(f"""
            SELECT COALESCE(category,'(unmapped)') AS k, COUNT(*) AS n,
                   ROUND(COALESCE(SUM(value_amount),0)) AS total_v
            FROM tenders {where}
            GROUP BY category ORDER BY n DESC""")
        by_source = q(f"""
            SELECT source AS k, COUNT(*) AS n
            FROM tenders {where}
            GROUP BY source ORDER BY n DESC""")
        by_value_band = q(f"""
            SELECT CASE
                WHEN value_amount IS NULL THEN 'Unknown'
                WHEN value_amount < 30000 THEN '< £30k'
                WHEN value_amount < 135000 THEN '£30k–£135k'
                WHEN value_amount < 300000 THEN '£135k–£300k'
                WHEN value_amount < 664000 THEN '£300k–£664k'
                ELSE '> £664k' END AS k,
                COUNT(*) AS n
            FROM tenders {where} GROUP BY k""")
        by_deadline = q(f"""
            SELECT CASE
                WHEN deadline IS NULL THEN 'No deadline'
                WHEN deadline <= date('now','+7 days') THEN 'This week'
                WHEN deadline <= date('now','+30 days') THEN 'This month'
                WHEN deadline <= date('now','+90 days') THEN '1–3 months'
                ELSE '3+ months' END AS k,
                COUNT(*) AS n
            FROM tenders {where} GROUP BY k""")
        # Basic counts + sum + closing-≤7d count (added Phase 2 for /dashboard KPI).
        # `open` uses is_open (heuristic, cheap) — accurate enough for the
        # summary KPI; the definitive "live" count lives at /api/live-stats.
        # `closing_7d` also compares deadline to now, catching stale rows
        # where is_open=1 but the deadline has passed.
        totals = conn.execute(
            f"SELECT COUNT(*), SUM(is_open), "
            f"       ROUND(COALESCE(SUM(value_amount),0)), "
            f"       SUM(CASE WHEN deadline IS NOT NULL "
            f"                 AND deadline >= datetime('now') "
            f"                 AND deadline <= datetime('now', '+7 days') "
            f"           THEN 1 ELSE 0 END) "
            f"FROM tenders {where}", params).fetchone()

        # Median + P90 — much more informative than the mean for tender values
        # (heavy-tailed distribution, a handful of £bn frameworks skew the mean).
        # Computed only over notices with a known value.
        median_row = conn.execute(f"""
            WITH ranked AS (
                SELECT value_amount, ROW_NUMBER() OVER (ORDER BY value_amount) rn,
                       COUNT(*) OVER () c
                FROM tenders {where} AND value_amount IS NOT NULL AND value_amount > 0
            )
            SELECT ROUND(AVG(value_amount)) FROM ranked
            WHERE rn IN ((c + 1) / 2, (c + 2) / 2)
        """ if where else f"""
            WITH ranked AS (
                SELECT value_amount, ROW_NUMBER() OVER (ORDER BY value_amount) rn,
                       COUNT(*) OVER () c
                FROM tenders WHERE value_amount IS NOT NULL AND value_amount > 0
            )
            SELECT ROUND(AVG(value_amount)) FROM ranked
            WHERE rn IN ((c + 1) / 2, (c + 2) / 2)
        """, params).fetchone()
        median_value = median_row[0] if median_row else 0

        # 90th percentile — the "big end" of the sensible range
        p90_row = conn.execute(f"""
            WITH ranked AS (
                SELECT value_amount, ROW_NUMBER() OVER (ORDER BY value_amount) rn,
                       COUNT(*) OVER () c
                FROM tenders {where} AND value_amount IS NOT NULL AND value_amount > 0
            )
            SELECT ROUND(value_amount) FROM ranked WHERE rn = CAST(0.9 * c AS INTEGER)
        """ if where else """
            WITH ranked AS (
                SELECT value_amount, ROW_NUMBER() OVER (ORDER BY value_amount) rn,
                       COUNT(*) OVER () c
                FROM tenders WHERE value_amount IS NOT NULL AND value_amount > 0
            )
            SELECT ROUND(value_amount) FROM ranked WHERE rn = CAST(0.9 * c AS INTEGER)
        """, params).fetchone()
        p90_value = p90_row[0] if p90_row else 0

        # Sweet-spot count: notices in £50k-£300k with a mapped category
        sweet_row = conn.execute(
            f"SELECT COUNT(*) FROM tenders {where} "
            f"AND value_amount BETWEEN 50000 AND 300000 "
            f"AND category IS NOT NULL"
            if where else
            "SELECT COUNT(*) FROM tenders "
            "WHERE value_amount BETWEEN 50000 AND 300000 AND category IS NOT NULL",
            params,
        ).fetchone()
        sweet_spot_count = sweet_row[0] if sweet_row else 0

        # Sum of values inside sensible SME range (<= £5M) — meaningful "opportunity value"
        in_range_row = conn.execute(
            f"SELECT ROUND(COALESCE(SUM(value_amount),0)) FROM tenders {where} "
            f"AND value_amount BETWEEN 0 AND 5000000"
            if where else
            "SELECT ROUND(COALESCE(SUM(value_amount),0)) FROM tenders "
            "WHERE value_amount BETWEEN 0 AND 5000000",
            params,
        ).fetchone()
        in_range_total = in_range_row[0] if in_range_row else 0
    finally:
        conn.close()

    def dictify(rows, keys=("k", "n")):
        return [dict(zip(keys, r)) for r in rows]

    return jsonify({
        "totals": {
            "notices": totals[0], "open": totals[1] or 0,
            "total_value": totals[2] or 0,
            "closing_7d": totals[3] or 0,     # Phase 2 addition (Dashboard KPI)
            "median_value": median_value or 0,
            "p90_value": p90_value or 0,
            "sweet_spot_count": sweet_spot_count,
            "in_range_total": in_range_total or 0,
        },
        "by_category": dictify(by_cat, ("k", "n", "total_v")),
        "by_source": dictify(by_source),
        "by_value_band": dictify(by_value_band),
        "by_deadline": dictify(by_deadline),
    })


# --- pivot data (raw denormalised rows, filtered) ------------------------ #
@app.route("/api/pivot")
def api_pivot():
    where, params = _build_where(request.args)
    conn = db.connect()
    try:
        cur = conn.execute(f"""
            SELECT source, COALESCE(category,'(unmapped)') AS category,
                   COALESCE(buyer_region,'(unspecified)') AS region,
                   COALESCE(country,'?') AS country,
                   value_amount,
                   substr(deadline,1,7) AS deadline_month,
                   CASE WHEN is_open=1 THEN 'Open' ELSE 'Closed' END AS status
            FROM tenders {where}""", params)
        cols = [d[0] for d in cur.description]
        rows = [list(r) for r in cur.fetchall()]
    finally:
        conn.close()
    return jsonify({"columns": cols, "rows": rows})


# --- scrape controls ----------------------------------------------------- #
@app.route("/api/scrape", methods=["POST"])
def api_scrape():
    global _scrape_state
    with _scrape_lock:
        if _scrape_state["running"]:
            return jsonify({"ok": False, "error": "A scrape is already running."}), 409
        payload = request.get_json(silent=True) or {}
        days_back = int(payload.get("days_back", 30))
        max_pages = int(payload.get("max_pages", 25))
        _scrape_state = {
            "running": True, "started_at": None, "finished_at": None,
            "days_back": days_back, "log": [], "error": None,
            "before": {}, "after": {},
        }
        threading.Thread(
            target=_run_scrape, args=(days_back, max_pages), daemon=True
        ).start()
    return jsonify({"ok": True, "days_back": days_back, "max_pages": max_pages})


@app.route("/api/scrape/status")
def api_scrape_status():
    return jsonify(_scrape_state)


# --- export -------------------------------------------------------------- #
@app.route("/api/export")
def api_export():
    fmt = request.args.get("format", "xlsx")
    stage = request.args.get("stage", "tender")
    where, params = _build_where(request.args, stage=stage)
    if stage == "award":
        cols = ("source,title,category,cpv_code,cpv_description,buyer_name,"
                "buyer_region,country,"
                "awarded_supplier_name,awarded_supplier_id,awarded_supplier_count,"
                "awarded_value_amount,awarded_value_currency,"
                "awarded_date,contract_start_date,contract_end_date,"
                "value_amount,notice_url,source_api_url,ocid")
    else:
        cols = ("source,title,category,cpv_code,cpv_description,buyer_name,"
                "buyer_region,country,value_amount,value_currency,"
                "published_date,deadline,status,is_open,notice_url,source_api_url")
    conn = db.connect()
    try:
        cur = conn.execute(
            f"SELECT {cols} FROM tenders {where} "
            f"ORDER BY (deadline IS NULL), deadline ASC",
            params,
        )
        col_names = [d[0] for d in cur.description]
        rows = cur.fetchall()
    finally:
        conn.close()

    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
    if fmt == "csv":
        buf = io.StringIO()
        import csv
        w = csv.writer(buf)
        w.writerow(col_names)
        w.writerows(rows)
        return Response(
            buf.getvalue(), mimetype="text/csv",
            headers={"Content-Disposition":
                     f'attachment; filename="bidgov-compass-{ts}.csv"'},
        )

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            abort(500, "openpyxl not installed on server.")
        wb = Workbook()
        ws = wb.active
        ws.title = "Tenders"
        ws.append(col_names)
        # header styling — GovBid palette
        header_fill = PatternFill("solid", fgColor="54565B")
        header_font = Font(name="Inter", bold=True, color="FFFFFF")
        for c in ws[1]:
            c.fill = header_fill; c.font = header_font
            c.alignment = Alignment(vertical="center")
        for r in rows:
            ws.append(list(r))
        ws.freeze_panes = "A2"
        for i, col in enumerate(col_names, 1):
            ws.column_dimensions[chr(64+i) if i < 27 else "A" + chr(64+i-26)].width = \
                max(12, min(40, len(col) + 4))
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"bidgov-compass-{ts}.xlsx",
        )

    abort(400, f"Unknown format: {fmt}")


# --- chat --------------------------------------------------------------- #
@app.route("/api/chat", methods=["POST"])
def api_chat():
    payload = request.get_json(silent=True) or {}
    question = (payload.get("question") or "").strip()
    if not question:
        return jsonify({"error": "Empty question"}), 400
    conn = db.connect()
    try:
        result = chat_mod.answer(conn, question)
    finally:
        conn.close()
    return jsonify(result)


@app.route("/health")
def health():
    return jsonify({"ok": True, "ts": datetime.now(timezone.utc).isoformat()})


if __name__ == "__main__":
    port = int(__import__("os").environ.get("PORT", "5057"))
    print(f"\nBidGov Compass -> http://127.0.0.1:{port}\n")
    app.run(host="127.0.0.1", port=port, debug=False, threaded=True)
